import json
import openpyxl

with open('data.txt', encoding='utf8') as f:
    data = f.readlines()

print(type(data),'\n=====\n')

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("ReportCalls.xlsx")
 
# Define variable to read sheet
dataframe1 = dataframe.active
 
# Iterate the loop to read the cell values
for row in range(1, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.min_column):
        print(col[row].value)


# import pandas lib as pd
import pandas as pd
 
# require_cols = [0, 3]
require_cols = [0]
 
# only read specific columns from an excel file
required_df = pd.read_excel('ReportCalls.xlsx', usecols=require_cols)
df = pd.DataFrame(required_df)

for i in df['Conference Id']:
    print(i)
